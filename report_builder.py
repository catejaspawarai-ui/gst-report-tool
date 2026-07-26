"""
Builds the consolidated GST report in the "months horizontal" layout used by
your reference Annual Report workbook:
    Company GSTIN | Section | Type | <Month 1> | <Month 2> | ... | Total
One row per (Section, Type) combination, one column per period, a live-formula
Total column at the end. Multiple GSTINs stack as additional row blocks in the
same sheet (their own GSTIN in column A), so this scales the same way whether
you're running one GSTIN or the full multi-GSTIN list from your dashboard.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT_NAME = "Calibri"
FONT_SIZE = 10

# No red for negatives -- parentheses only, per house style.
INR_FMT = '₹#,##,##0.00;(₹#,##,##0.00);"-"'

HEADER_FILL = PatternFill(start_color="3A3630", end_color="3A3630", fill_type="solid")  # charcoal
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="F2ECE3", size=FONT_SIZE)  # warm cream text
BODY_FONT = Font(name=FONT_NAME, size=FONT_SIZE)
GSTIN_FONT = Font(name=FONT_NAME, size=FONT_SIZE, bold=True)
SECTION_FILL = PatternFill(start_color="EDE9E1", end_color="EDE9E1", fill_type="solid")  # warm neutral tint
DIFF_FILL = PatternFill(start_color="F3EAD3", end_color="F3EAD3", fill_type="solid")  # light beige -- Difference rows only
COUNT_FMT = '#,##0;(#,##0);"-"'  # plain number, no currency symbol -- for document counts
DIFF_FONT = Font(name=FONT_NAME, size=FONT_SIZE, italic=True)
TOTAL_FONT = Font(name=FONT_NAME, size=FONT_SIZE, bold=True)
THIN_BORDER = Border(*(Side(style="thin", color="D9D9D9"),) * 4)

MONTHS = ["", "Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

TYPE_LABELS = {"txval": "Taxable Value", "igst": "Integrated Tax", "cgst": "Central Tax", "sgst": "State/UT Tax"}


def period_sort_key(period):
    """MMYYYY -> YYYYMM as int, for chronological sorting."""
    mm, yyyy = int(period[:2]), int(period[2:])
    return yyyy * 100 + mm


def period_label(period):
    mm, yyyy = int(period[:2]), period[2:]
    return f"{MONTHS[mm]} {yyyy}"


def sorted_periods(*record_lists):
    periods = set()
    for records in record_lists:
        for r in records:
            if r.get("period"):
                periods.add(r["period"])
    return sorted(periods, key=period_sort_key)


def _write_header(ws, period_cols, row=1):
    headers = ["Company GSTIN", "Section", "Type"] + [period_label(p) for p in period_cols] + ["Total"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = THIN_BORDER
    ws.freeze_panes = ws.cell(row=row + 1, column=4)
    ws.auto_filter.ref = ws.dimensions
    return len(headers)


def _write_section_banner(ws, row, text, num_cols):
    c = ws.cell(row=row, column=2, value=text)
    c.font = Font(name=FONT_NAME, bold=True, size=FONT_SIZE, color="00707A")  # dark teal, readable on the tint
    for col in range(1, num_cols + 1):
        ws.cell(row=row, column=col).fill = SECTION_FILL
        ws.cell(row=row, column=col).border = THIN_BORDER


def _write_data_row(ws, row, gstin, section, type_name, values, num_cols, italic=False, bold_label=False, flag=None, number_format=None, text_only=False):
    fmt = number_format or INR_FMT
    font = DIFF_FONT if italic else BODY_FONT
    label_font = TOTAL_FONT if bold_label else font
    ws.cell(row=row, column=1, value=gstin).font = font
    ws.cell(row=row, column=2, value=section).font = label_font
    ws.cell(row=row, column=3, value=type_name).font = font
    first_period_col = 4
    for i, v in enumerate(values):
        col = first_period_col + i
        c = ws.cell(row=row, column=col, value=v)
        if v is not None and not text_only:
            c.number_format = fmt
        c.font = font
        c.border = THIN_BORDER
    total_col = first_period_col + len(values)
    if text_only:
        # Dates/text can't be meaningfully summed -- leave the Total column blank instead
        tc = ws.cell(row=row, column=total_col, value=None)
        tc.font = TOTAL_FONT if not italic else font
    else:
        first_letter = get_column_letter(first_period_col)
        last_letter = get_column_letter(total_col - 1)
        tc = ws.cell(row=row, column=total_col, value=f"=SUM({first_letter}{row}:{last_letter}{row})")
        tc.number_format = fmt
        tc.font = TOTAL_FONT if not italic else font
    for col in range(1, num_cols + 1):
        ws.cell(row=row, column=col).border = THIN_BORDER
    if flag and flag != "OK":
        for col in range(1, num_cols + 1):
            ws.cell(row=row, column=col).fill = DIFF_FILL


def _autofit(ws, num_period_cols):
    ws.column_dimensions["A"].width = 17
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 16
    for i in range(num_period_cols + 1):
        ws.column_dimensions[get_column_letter(4 + i)].width = 15


def build_overview_sheet(wb, company_name, gstin, financial_year, data_notes=None):
    ws = wb.create_sheet("Overview")
    ws.sheet_properties.tabColor = "FFFFFF"  # explicit white -- overrides Excel theme auto-tinting inactive tabs
    rows = [
        ("Company Name", company_name or ""),
        ("Contents", "GST Returns Report"),
        ("GSTIN", gstin or ""),
        ("Period", financial_year or ""),
    ]
    ws.cell(row=1, column=2, value="Report Details").font = Font(name=FONT_NAME, bold=True, size=12)
    row_i = 2
    for label, value in rows:
        ws.cell(row=row_i, column=2, value=label).font = Font(name=FONT_NAME, bold=True, size=FONT_SIZE)
        ws.cell(row=row_i, column=3, value=value).font = BODY_FONT
        row_i += 1

    if data_notes:
        row_i += 1
        ws.cell(row=row_i, column=2, value="Data Notes").font = Font(name=FONT_NAME, bold=True, size=12)
        row_i += 1
        for note in data_notes:
            c = ws.cell(row=row_i, column=2, value=f"\u2022 {note}")
            c.font = BODY_FONT
            row_i += 1

    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 35


def _records_by_gstin_period(records):
    out = {}
    for r in records:
        out.setdefault(r["gstin"], {})[r["period"]] = r
    return out


def build_gstr3b_sheet(wb, records, periods):
    ws = wb.create_sheet("GSTR-3B")
    ws.sheet_properties.tabColor = "FFFFFF"  # explicit white -- overrides Excel theme auto-tinting inactive tabs
    num_cols = _write_header(ws, periods)
    row = 2
    by_gstin = _records_by_gstin_period(records)

    # Each group is (banner text or None, [(section, type, extractor), ...])
    # banner=None means "always show this banner row, even with no data beneath it"
    # (matches the official GSTR-3B table, which shows empty sub-tables too).
    groups = [
        ("3.1 Details of Outward Supplies and inward supplies liable to reverse charge", [
            ("3.1.A Outward taxable supplies (excl. zero rated)", "Supply Value", lambda r: r["outward_taxable_value"]["txval"]),
            ("3.1.A Outward taxable supplies (excl. zero rated)", "Integrated Tax", lambda r: r["outward_taxable_value"]["igst"]),
            ("3.1.A Outward taxable supplies (excl. zero rated)", "Central Tax", lambda r: r["outward_taxable_value"]["cgst"]),
            ("3.1.A Outward taxable supplies (excl. zero rated)", "State/UT Tax", lambda r: r["outward_taxable_value"]["sgst"]),
            ("3.1.B Outward taxable supplies (zero rated)", "Supply Value", lambda r: r["zero_rated"]["txval"]),
            ("3.1.B Outward taxable supplies (zero rated)", "Integrated Tax", lambda r: r["zero_rated"]["igst"]),
            ("3.1.C Nil rated / exempt outward supplies", "Supply Value", lambda r: r["nil_exempt"]["txval"]),
            ("3.1.D Inward supplies liable to reverse charge", "Supply Value", lambda r: r["rcm_inward_liability"]["txval"]),
            ("3.1.D Inward supplies liable to reverse charge", "Integrated Tax", lambda r: r["rcm_inward_liability"]["igst"]),
            ("3.1.D Inward supplies liable to reverse charge", "Central Tax", lambda r: r["rcm_inward_liability"]["cgst"]),
            ("3.1.D Inward supplies liable to reverse charge", "State/UT Tax", lambda r: r["rcm_inward_liability"]["sgst"]),
            ("3.1.E Non-GST outward supplies", "Supply Value", lambda r: r["non_gst_outward_txval"]),
        ]),
        ("3.1.1 Details of Supplies notified under section 9(5) of the CGST Act, 2017", [
            ("3.1.1 Section 9(5) supplies (e-commerce operator pays tax)", "Supply Value", lambda r: r["eco_supplies_9_5"]["txval"]),
            ("3.1.1 Section 9(5) supplies (e-commerce operator pays tax)", "Integrated Tax", lambda r: r["eco_supplies_9_5"]["igst"]),
            ("3.1.1 Section 9(5) supplies (e-commerce operator pays tax)", "Central Tax", lambda r: r["eco_supplies_9_5"]["cgst"]),
            ("3.1.1 Section 9(5) supplies (e-commerce operator pays tax)", "State/UT Tax", lambda r: r["eco_supplies_9_5"]["sgst"]),
        ]),
        ("4(A) ITC Available", [
            ("4.A.1 Import of goods", "Integrated Tax", lambda r: r["itc_by_type"].get("IMPG", {}).get("igst", 0)),
            ("4.A.2 Import of services", "Integrated Tax", lambda r: r["itc_by_type"].get("IMPS", {}).get("igst", 0)),
            ("4.A.3 Inward supplies liable to reverse charge", "Integrated Tax", lambda r: r["itc_by_type"].get("ISRC", {}).get("igst", 0)),
            ("4.A.3 Inward supplies liable to reverse charge", "Central Tax", lambda r: r["itc_by_type"].get("ISRC", {}).get("cgst", 0)),
            ("4.A.3 Inward supplies liable to reverse charge", "State/UT Tax", lambda r: r["itc_by_type"].get("ISRC", {}).get("sgst", 0)),
            ("4.A.4 Inward supplies from ISD", "Integrated Tax", lambda r: r["itc_by_type"].get("ISD", {}).get("igst", 0)),
            ("4.A.4 Inward supplies from ISD", "Central Tax", lambda r: r["itc_by_type"].get("ISD", {}).get("cgst", 0)),
            ("4.A.4 Inward supplies from ISD", "State/UT Tax", lambda r: r["itc_by_type"].get("ISD", {}).get("sgst", 0)),
            ("4.A.5 All other ITC", "Integrated Tax", lambda r: r["itc_by_type"].get("OTH", {}).get("igst", 0)),
            ("4.A.5 All other ITC", "Central Tax", lambda r: r["itc_by_type"].get("OTH", {}).get("cgst", 0)),
            ("4.A.5 All other ITC", "State/UT Tax", lambda r: r["itc_by_type"].get("OTH", {}).get("sgst", 0)),
            ("4.A.5 All other ITC", "Cess", lambda r: r["itc_by_type"].get("OTH", {}).get("cess", 0)),
        ]),
        ("4(B) ITC Reversed", [
            ("4.B.1 As per Rule 42/43 etc.", "Integrated Tax", lambda r: -r["itc_reversed_by_type"].get("RUL", {}).get("igst", 0)),
            ("4.B.1 As per Rule 42/43 etc.", "Central Tax", lambda r: -r["itc_reversed_by_type"].get("RUL", {}).get("cgst", 0)),
            ("4.B.1 As per Rule 42/43 etc.", "State/UT Tax", lambda r: -r["itc_reversed_by_type"].get("RUL", {}).get("sgst", 0)),
            ("4.B.2 Others", "Integrated Tax", lambda r: -r["itc_reversed_by_type"].get("OTH", {}).get("igst", 0)),
            ("4.B.2 Others", "Central Tax", lambda r: -r["itc_reversed_by_type"].get("OTH", {}).get("cgst", 0)),
            ("4.B.2 Others", "State/UT Tax", lambda r: -r["itc_reversed_by_type"].get("OTH", {}).get("sgst", 0)),
            ("4.B.2 Others", "Cess", lambda r: -r["itc_reversed_by_type"].get("OTH", {}).get("cess", 0)),
        ]),
        ("4(C) Net ITC Available (A - B)", [
            ("Net ITC", "Integrated Tax", lambda r: r["itc_net"]["igst"]),
            ("Net ITC", "Central Tax", lambda r: r["itc_net"]["cgst"]),
            ("Net ITC", "State/UT Tax", lambda r: r["itc_net"]["sgst"]),
            ("Net ITC", "Cess", lambda r: r["itc_net"]["cess"]),
        ]),
        # NOTE: the raw JSON's 'RUL'/'OTH' type codes mean different things in
        # itc_inelg than they do in itc_rev, despite sharing the same codes --
        # confirmed by exact numeric match against a real filed return:
        #   itc_inelg[RUL] = 4.D.1 ITC RECLAIMED (not "ineligible" at all --
        #     this is credit reversed earlier under Table 4(B)(2) being added
        #     back once the underlying condition is met, e.g. payment made
        #     within 180 days)
        #   itc_inelg[OTH] = 4.D.2 genuinely ineligible ITC (Section 16(4)
        #     time-barred claims, and place-of-supply restrictions)
        ("4(D) Other Details", [
            ("4.D.1 ITC Reclaimed (reversed earlier under Table 4(B)(2))", "Integrated Tax", lambda r: r["itc_ineligible_by_type"].get("RUL", {}).get("igst", 0)),
            ("4.D.1 ITC Reclaimed (reversed earlier under Table 4(B)(2))", "Central Tax", lambda r: r["itc_ineligible_by_type"].get("RUL", {}).get("cgst", 0)),
            ("4.D.1 ITC Reclaimed (reversed earlier under Table 4(B)(2))", "State/UT Tax", lambda r: r["itc_ineligible_by_type"].get("RUL", {}).get("sgst", 0)),
            ("4.D.2 Ineligible ITC (Section 16(4) & PoS restrictions)", "Integrated Tax", lambda r: r["itc_ineligible_by_type"].get("OTH", {}).get("igst", 0)),
            ("4.D.2 Ineligible ITC (Section 16(4) & PoS restrictions)", "Central Tax", lambda r: r["itc_ineligible_by_type"].get("OTH", {}).get("cgst", 0)),
            ("4.D.2 Ineligible ITC (Section 16(4) & PoS restrictions)", "State/UT Tax", lambda r: r["itc_ineligible_by_type"].get("OTH", {}).get("sgst", 0)),
        ]),
        ("Total Liability & Payment Summary", [
            ("Total Liability (Other than reverse charge)", "Value", lambda r: r["total_liability_tax"]),
            ("Total Liability (Reverse Charge)", "Value", lambda r: sum(r["rcm_inward_liability"][h] for h in ("igst", "cgst", "sgst", "cess"))),
            ("Paid using ITC", "Value", lambda r: r["tax_paid_itc"]),
            ("Paid using Cash", "Value", lambda r: r["tax_paid_cash"]),
        ]),
        ("Payment of Tax", [
            ("Integrated Tax", "Integrated Tax ITC", lambda r: r["payment_matrix"]["IGST"]["via_igst_itc"]),
            ("Integrated Tax", "Central Tax ITC", lambda r: r["payment_matrix"]["IGST"]["via_cgst_itc"]),
            ("Integrated Tax", "State/UT Tax ITC", lambda r: r["payment_matrix"]["IGST"]["via_sgst_itc"]),
            ("Integrated Tax", "Cash", lambda r: r["payment_matrix"]["IGST"]["via_cash"]),
            ("Central Tax", "Integrated Tax ITC", lambda r: r["payment_matrix"]["CGST"]["via_igst_itc"]),
            ("Central Tax", "Central Tax ITC", lambda r: r["payment_matrix"]["CGST"]["via_cgst_itc"]),
            ("Central Tax", "Cash", lambda r: r["payment_matrix"]["CGST"]["via_cash"]),
            ("State/UT Tax", "Integrated Tax ITC", lambda r: r["payment_matrix"]["SGST"]["via_igst_itc"]),
            ("State/UT Tax", "State/UT Tax ITC", lambda r: r["payment_matrix"]["SGST"]["via_sgst_itc"]),
            ("State/UT Tax", "Cash", lambda r: r["payment_matrix"]["SGST"]["via_cash"]),
            ("Cess", "Cess ITC", lambda r: r["payment_matrix"]["CESS"]["via_cess_itc"]),
            ("Cess", "Cash", lambda r: r["payment_matrix"]["CESS"]["via_cash"]),
        ]),
        ("Interest & Late Fee", [
            ("5.1 Interest", "Integrated Tax", lambda r: r["interest_due"]["igst"]),
            ("5.1 Interest", "Central Tax", lambda r: r["interest_due"]["cgst"]),
            ("5.1 Interest", "State/UT Tax", lambda r: r["interest_due"]["sgst"]),
            ("6.1 Late Fee", "Central Tax", lambda r: r["late_fee_due"]["cgst"]),
            ("6.1 Late Fee", "State/UT Tax", lambda r: r["late_fee_due"]["sgst"]),
        ]),
    ]

    for gstin, by_period in by_gstin.items():
        for banner, row_defs in groups:
            _write_section_banner(ws, row, banner, num_cols)
            row += 1
            for section, type_name, fn in row_defs:
                values = [fn(by_period[p]) if p in by_period else None for p in periods]
                if all(v in (None, 0) for v in values):
                    continue
                _write_data_row(ws, row, gstin, section, type_name, values, num_cols)
                row += 1

    _autofit(ws, len(periods))


def build_gstr1_sheet(wb, records, periods):
    ws = wb.create_sheet("GSTR-1")
    ws.sheet_properties.tabColor = "FFFFFF"  # explicit white -- overrides Excel theme auto-tinting inactive tabs
    num_cols = _write_header(ws, periods)
    row = 2
    by_gstin = _records_by_gstin_period(records)

    row_defs = [
        ("B2B Invoices", "Taxable Value", lambda r: r["b2b"]["txval"]),
        ("B2B Invoices", "Integrated Tax", lambda r: r["b2b"]["igst"]),
        ("B2B Invoices", "Central Tax", lambda r: r["b2b"]["cgst"]),
        ("B2B Invoices", "State/UT Tax", lambda r: r["b2b"]["sgst"]),
        ("B2B Credit Notes", "Taxable Value", lambda r: -r["cdnr_credit"]["txval"]),
        ("B2B Credit Notes", "Integrated Tax", lambda r: -r["cdnr_credit"]["igst"]),
        ("B2B Credit Notes", "Central Tax", lambda r: -r["cdnr_credit"]["cgst"]),
        ("B2B Credit Notes", "State/UT Tax", lambda r: -r["cdnr_credit"]["sgst"]),
        ("B2B Debit Notes", "Taxable Value", lambda r: r["cdnr_debit"]["txval"]),
        ("B2B Debit Notes", "Integrated Tax", lambda r: r["cdnr_debit"]["igst"]),
        ("EXP-WP Invoices", "Taxable Value", lambda r: r["exp_wpay"]["txval"]),
        ("EXP-WP Invoices", "Integrated Tax", lambda r: r["exp_wpay"]["igst"]),
        ("EXP-WOP Invoices", "Taxable Value", lambda r: r["exp_wopay"]["txval"]),
        ("Unregistered Notes (domestic)", "Taxable Value", lambda r: r["cdnur_debit"]["txval"] - r["cdnur_credit"]["txval"]),
        ("Unregistered Notes (export)", "Taxable Value", lambda r: r["cdnur_exp_debit"]["txval"] - r["cdnur_exp_credit"]["txval"]),
        ("B2CS", "Taxable Value", lambda r: r["b2cs"]["txval"]),
        ("SEZ Supplies (with payment)", "Taxable Value", lambda r: r["b2b_sez_wp"]["txval"]),
        ("SEZ Supplies (with payment)", "Integrated Tax", lambda r: r["b2b_sez_wp"]["igst"]),
        ("SEZ Supplies (without payment)", "Taxable Value", lambda r: r["b2b_sez_wop"]["txval"]),
        ("Deemed Export", "Taxable Value", lambda r: r["b2b_deemed_exp"]["txval"]),
        ("Deemed Export", "Integrated Tax", lambda r: r["b2b_deemed_exp"]["igst"]),
        ("Deemed Export", "Central Tax", lambda r: r["b2b_deemed_exp"]["cgst"]),
        ("Deemed Export", "State/UT Tax", lambda r: r["b2b_deemed_exp"]["sgst"]),
        ("CDNR Amendments (registered, credit)", "Taxable Value", lambda r: -r["cdnra_credit"]["txval"]),
        ("CDNR Amendments (registered, debit)", "Taxable Value", lambda r: r["cdnra_debit"]["txval"]),
        ("E-commerce U/s 9(5) - B2B", "Taxable Value", lambda r: r["ecom_b2b"]["txval"]),
        ("E-commerce U/s 9(5) - B2B", "Integrated Tax", lambda r: r["ecom_b2b"]["igst"]),
        ("E-commerce U/s 9(5) - B2C", "Taxable Value", lambda r: r["ecom_b2c"]["txval"]),
        ("E-commerce U/s 9(5) - B2C", "Integrated Tax", lambda r: r["ecom_b2c"]["igst"]),
        ("E-commerce U/s 9(5) - URP2C", "Taxable Value", lambda r: r["ecom_urp2c"]["txval"]),
        ("E-commerce U/s 9(5) - URP2C", "Integrated Tax", lambda r: r["ecom_urp2c"]["igst"]),
        ("B2CS", "Integrated Tax", lambda r: r["b2cs"]["igst"]),
        ("B2CS", "Central Tax", lambda r: r["b2cs"]["cgst"]),
        ("B2CS", "State/UT Tax", lambda r: r["b2cs"]["sgst"]),
        ("B2B Amendments", "Taxable Value", lambda r: r["b2b_amendments"]["txval"]),
        ("B2B Amendments", "Integrated Tax", lambda r: r["b2b_amendments"]["igst"]),
        ("B2B Amendments", "Central Tax", lambda r: r["b2b_amendments"]["cgst"]),
        ("B2B Amendments", "State/UT Tax", lambda r: r["b2b_amendments"]["sgst"]),
        ("Export Amendments (WP)", "Taxable Value", lambda r: r["exp_wpay_amendments"]["txval"]),
        ("Export Amendments (WP)", "Integrated Tax", lambda r: r["exp_wpay_amendments"]["igst"]),
        ("Export Amendments (WOP)", "Taxable Value", lambda r: r["exp_wopay_amendments"]["txval"]),
        ("B2C Large", "Taxable Value", lambda r: r["b2cl"]["txval"]),
        ("B2C Large", "Integrated Tax", lambda r: r["b2cl"]["igst"]),
        ("B2C Large Amendments", "Taxable Value", lambda r: r["b2cl_amendments"]["txval"]),
        ("B2C Large Amendments", "Integrated Tax", lambda r: r["b2cl_amendments"]["igst"]),
        ("B2CS Amendments", "Taxable Value", lambda r: r["b2cs_amendments"]["txval"]),
        ("B2CS Amendments", "Integrated Tax", lambda r: r["b2cs_amendments"]["igst"]),
        ("B2CS Amendments", "Central Tax", lambda r: r["b2cs_amendments"]["cgst"]),
        ("B2CS Amendments", "State/UT Tax", lambda r: r["b2cs_amendments"]["sgst"]),
        ("Unregistered Notes Amendments (credit)", "Taxable Value", lambda r: -r["cdnura_credit"]["txval"]),
        ("Unregistered Notes Amendments (debit)", "Taxable Value", lambda r: r["cdnura_debit"]["txval"]),
        ("Nil Rated Supplies", "Exempted", lambda r: r["nil_supplies"]["exempted"]),
        ("Nil Rated Supplies", "Nil Rated", lambda r: r["nil_supplies"]["nil_rated"]),
        ("Nil Rated Supplies", "Non-GST", lambda r: r["nil_supplies"]["non_gst"]),
        ("11A Advances Received", "Advance Amount", lambda r: r["advances_received"]["ad_amt"]),
        ("11A Advances Received", "Integrated Tax", lambda r: r["advances_received"]["igst"]),
        ("11A Advances Received", "Central Tax", lambda r: r["advances_received"]["cgst"]),
        ("11A Advances Received", "State/UT Tax", lambda r: r["advances_received"]["sgst"]),
        ("11A Advances Received Amendments", "Advance Amount", lambda r: r["advances_received_amendments"]["ad_amt"]),
        ("11B Advances Adjusted", "Advance Amount", lambda r: r["advances_adjusted"]["ad_amt"]),
        ("11B Advances Adjusted", "Integrated Tax", lambda r: r["advances_adjusted"]["igst"]),
        ("11B Advances Adjusted", "Central Tax", lambda r: r["advances_adjusted"]["cgst"]),
        ("11B Advances Adjusted", "State/UT Tax", lambda r: r["advances_adjusted"]["sgst"]),
        ("11B Advances Adjusted Amendments", "Advance Amount", lambda r: r["advances_adjusted_amendments"]["ad_amt"]),
    ]

    for gstin, by_period in by_gstin.items():
        for section, type_name, fn in row_defs:
            values = [fn(by_period[p]) if p in by_period else None for p in periods]
            if all(v in (None, 0) for v in values):
                continue
            _write_data_row(ws, row, gstin, section, type_name, values, num_cols)
            row += 1

    _autofit(ws, len(periods))


def build_hsn_summary_sheet(wb, gstr1_records, periods):
    """HSN-wise summary (GSTR-1 Table 12), combining B2B + B2C entries.
    Row set is the UNION of every distinct (HSN code, description, rate) seen
    across all periods for a GSTIN -- not every period will have every HSN
    code, since product mix can change month to month."""
    ws = wb.create_sheet("HSN Summary")
    ws.sheet_properties.tabColor = "FFFFFF"
    num_cols = _write_header(ws, periods)
    row = 2
    by_gstin = _records_by_gstin_period(gstr1_records)

    for gstin, by_period in by_gstin.items():
        hsn_lookup = {}  # (period, key) -> hsn row dict
        hsn_keys = set()
        for p, r in by_period.items():
            for hsn_row in r.get("hsn_summary", []):
                key = (hsn_row["hsn_code"], hsn_row["description"], hsn_row["rate"])
                hsn_keys.add(key)
                hsn_lookup[(p, key)] = hsn_row

        for key in sorted(hsn_keys):
            hsn_code, desc, rate = key
            short_desc = (desc[:40] + "...") if len(desc) > 40 else desc
            section_label = f"HSN {hsn_code} - {short_desc} ({rate}%)"
            for type_name, field in [("Taxable Value", "txval"), ("Integrated Tax", "igst"),
                                       ("Central Tax", "cgst"), ("State/UT Tax", "sgst"), ("Cess", "cess")]:
                values = [hsn_lookup.get((p, key), {}).get(field) for p in periods]
                if all(v in (None, 0) for v in values):
                    continue
                _write_data_row(ws, row, gstin, section_label, type_name, values, num_cols)
                row += 1

    _autofit(ws, len(periods))


def build_doc_issue_summary_sheet(wb, gstr1_records, periods):
    """Document Issued summary (GSTR-1 Table 13) -- invoice/credit note/debit
    note/etc. series continuity and cancellation counts."""
    ws = wb.create_sheet("Doc Issued Summary")
    ws.sheet_properties.tabColor = "FFFFFF"
    num_cols = _write_header(ws, periods)
    row = 2
    by_gstin = _records_by_gstin_period(gstr1_records)

    for gstin, by_period in by_gstin.items():
        doc_lookup = {}
        doc_types = set()
        for p, r in by_period.items():
            for doc_row in r.get("doc_issue_summary", []):
                doc_types.add(doc_row["doc_type"])
                doc_lookup[(p, doc_row["doc_type"])] = doc_row

        for doc_type in sorted(doc_types):
            for type_name, field in [("Total Issued", "total_issued"), ("Cancelled", "total_cancelled"),
                                       ("Net Issued", "net_issued")]:
                values = [doc_lookup.get((p, doc_type), {}).get(field) for p in periods]
                if all(v in (None, 0) for v in values):
                    continue
                _write_data_row(ws, row, gstin, doc_type, type_name, values, num_cols, number_format=COUNT_FMT)
                row += 1

    _autofit(ws, len(periods))


def build_gstr2b_sheet(wb, detail_records, periods):
    """Built from GSTR-2B invoice-level detail files (File1/File2/...), not the
    Summary.json rollup -- gives the eligible/ineligible and regular/reverse-
    charge split within each category, verified against Summary.json's totals
    (exact match on ITC Not Available and Rejected across cross-checks)."""
    ws = wb.create_sheet("GSTR-2B")
    ws.sheet_properties.tabColor = "FFFFFF"
    num_cols = _write_header(ws, periods)
    row = 2
    by_gstin = _records_by_gstin_period(detail_records)

    from parsers.parse_r2b_detail import CATEGORY_LABELS
    elig_labels = {"eligible_regular": "Eligible (Regular)", "eligible_rcm": "Eligible (Reverse Charge)", "ineligible": "Ineligible"}

    row_defs = []
    for cat_key, cat_label in CATEGORY_LABELS.items():
        for elig_key, elig_label in elig_labels.items():
            section = f"{cat_label} - {elig_label}"
            for head, type_name in [("txval", "Taxable Value"), ("igst", "Integrated Tax"),
                                      ("cgst", "Central Tax"), ("sgst", "State/UT Tax"), ("cess", "Cess")]:
                row_defs.append((section, type_name,
                                  (lambda r, c=cat_key, e=elig_key, h=head: r["by_category"].get(c, {}).get(e, {}).get(h, 0))))

    for head, type_name in [("txval", "Taxable Value"), ("igst", "Integrated Tax"), ("cgst", "Central Tax"),
                             ("sgst", "State/UT Tax"), ("cess", "Cess")]:
        row_defs.append(("Import of Goods", type_name, (lambda r, h=head: r["imports"].get(h, 0))))
    for head, type_name in [("txval", "Taxable Value"), ("igst", "Integrated Tax"), ("cgst", "Central Tax"),
                             ("sgst", "State/UT Tax"), ("cess", "Cess")]:
        row_defs.append(("Rejected (IMS)", type_name, (lambda r, h=head: r["rejected"].get(h, 0))))

    for gstin, by_period in by_gstin.items():
        for section, type_name, fn in row_defs:
            values = [fn(by_period[p]) if p in by_period else None for p in periods]
            if all(v in (None, 0) for v in values):
                continue
            _write_data_row(ws, row, gstin, section, type_name, values, num_cols)
            row += 1

    _autofit(ws, len(periods))


# --- GSTR-3B vs GSTR-1: component-level detail, matching the reference layout ---
# Each R3B line is compared against every R1 component that feeds into it,
# with a Difference row underneath -- not a single blended R1 total.
_3B_VS_1_SECTIONS = [
    {
        "banner": "Outward taxable supplies (other than zero rated)",
        "r3b_label": "R3B: 3.1.A Outward taxable supplies (excluding zero rated)",
        "r3b_key": "outward_taxable_value",
        "components": [
            ("R1: B2B Invoices", "b2b", 1),
            ("R1: B2B Credit Notes", "cdnr_credit", -1),
            ("R1: B2B Debit Notes", "cdnr_debit", 1),
            ("R1: B2CS", "b2cs", 1),
            ("R1: Unregistered Debit Notes (domestic)", "cdnur_debit", 1),
            ("R1: Unregistered Credit Notes (domestic)", "cdnur_credit", -1),
            ("R1: B2B Amendments", "b2b_amendments", 1),
            ("R1: B2C Large", "b2cl", 1),
            ("R1: B2C Large Amendments", "b2cl_amendments", 1),
            ("R1: B2CS Amendments", "b2cs_amendments", 1),
            ("R1: Unregistered Notes Amendments (debit)", "cdnura_debit", 1),
            ("R1: Unregistered Notes Amendments (credit)", "cdnura_credit", -1),
            ("R1: CDNR Amendments (registered, debit)", "cdnra_debit", 1),
            ("R1: CDNR Amendments (registered, credit)", "cdnra_credit", -1),
            ("R1: 11A Advances Received", "advances_received", 1),
            ("R1: 11A Advances Received Amendments", "advances_received_amendments", 1),
            ("R1: 11B Advances Adjusted", "advances_adjusted", -1),
            ("R1: 11B Advances Adjusted Amendments", "advances_adjusted_amendments", -1),
            ("R1: Deemed Export", "b2b_deemed_exp", 1),
            ("R1: E-commerce U/s 9(5) - B2B", "ecom_b2b", 1),
            ("R1: E-commerce U/s 9(5) - B2C", "ecom_b2c", 1),
            ("R1: E-commerce U/s 9(5) - URP2C", "ecom_urp2c", 1),
        ],
        "types": ["txval", "igst", "cgst", "sgst"],
    },
    {
        "banner": "Outward taxable supplies (zero rated / exports)",
        "r3b_label": "R3B: 3.1.B Outward taxable supplies (zero rated)",
        "r3b_key": "zero_rated",
        "components": [
            ("R1: EXP-WP Invoices", "exp_wpay", 1),
            ("R1: EXP-WOP Invoices", "exp_wopay", 1),
            ("R1: Unregistered Debit Notes (export)", "cdnur_exp_debit", 1),
            ("R1: Unregistered Credit Notes (export)", "cdnur_exp_credit", -1),
            ("R1: Export Amendments (WP)", "exp_wpay_amendments", 1),
            ("R1: Export Amendments (WOP)", "exp_wopay_amendments", 1),
            ("R1: SEZ Supplies (with payment)", "b2b_sez_wp", 1),
            ("R1: SEZ Supplies (without payment)", "b2b_sez_wop", 1),
            ("R1: SEZ Credit/Debit Notes (debit)", "cdnr_sez_debit", 1),
            ("R1: SEZ Credit/Debit Notes (credit)", "cdnr_sez_credit", -1),
        ],
        "types": ["txval", "igst"],
    },
]


def build_3b_vs_1_sheet(wb, gstr3b_records, gstr1_records, periods):
    ws = wb.create_sheet("GSTR-3B vs GSTR-1")
    ws.sheet_properties.tabColor = "FFFFFF"  # explicit white -- overrides Excel theme auto-tinting inactive tabs
    num_cols = _write_header(ws, periods)
    row = 2

    r3b_by_gstin = _records_by_gstin_period(gstr3b_records)
    r1_by_gstin = _records_by_gstin_period(gstr1_records)
    gstins = sorted(set(r3b_by_gstin) | set(r1_by_gstin))

    for gstin in gstins:
        r3b_periods = r3b_by_gstin.get(gstin, {})
        r1_periods = r1_by_gstin.get(gstin, {})

        for section in _3B_VS_1_SECTIONS:
            _write_section_banner(ws, row, section["banner"], num_cols)
            row += 1
            for t in section["types"]:
                r3b_vals = [
                    r3b_periods[p][section["r3b_key"]].get(t, 0) if p in r3b_periods else None
                    for p in periods
                ]
                if not all(v in (None, 0) for v in r3b_vals):
                    _write_data_row(ws, row, gstin, section["r3b_label"], TYPE_LABELS[t], r3b_vals, num_cols)
                    row += 1

                component_value_lists = []
                for label, field, sign in section["components"]:
                    vals = [
                        sign * r1_periods[p][field].get(t, 0) if p in r1_periods else None
                        for p in periods
                    ]
                    if all(v in (None, 0) for v in vals):
                        continue
                    _write_data_row(ws, row, gstin, label, TYPE_LABELS[t], vals, num_cols)
                    row += 1
                    component_value_lists.append(vals)

                # Difference = R3B - sum of all R1 components, per period
                diff_vals = []
                any_data = False
                for i, p in enumerate(periods):
                    r3b_v = r3b_vals[i]
                    r1_sum = sum(v[i] or 0 for v in component_value_lists if v[i] is not None)
                    if r3b_v is None and not any(v[i] is not None for v in component_value_lists):
                        diff_vals.append(None)
                    else:
                        diff_vals.append(round((r3b_v or 0) - r1_sum, 2))
                        any_data = True
                if any_data:
                    max_abs_diff = max((abs(v) for v in diff_vals if v is not None), default=0)
                    flag = "REVIEW" if max_abs_diff > 100 else "OK"
                    _write_data_row(ws, row, gstin, "Difference", TYPE_LABELS[t], diff_vals, num_cols,
                                     italic=True, bold_label=True, flag=flag)
                    row += 1

    _autofit(ws, len(periods))


def build_3b_vs_2b_sheet(wb, comparisons, periods):
    ws = wb.create_sheet("GSTR-3B vs GSTR-2B")
    ws.sheet_properties.tabColor = "FFFFFF"  # explicit white -- overrides Excel theme auto-tinting inactive tabs
    num_cols = _write_header(ws, periods)
    row = 2
    by_gstin = {}
    for c in comparisons:
        by_gstin.setdefault(c["gstin"], {})[c["period"]] = c

    categories = ["Import of Goods", "Reverse Charge (RCM)", "ISD Credit", "All Other ITC (B2B + CDNR adj.)"]
    heads = ["IGST", "CGST", "SGST"]

    for gstin, by_period in by_gstin.items():
        for cat in categories:
            _write_section_banner(ws, row, f"ITC Claim - {cat}", num_cols)
            row += 1
            for head in heads:
                claimed, avail, flags = [], [], []
                for p in periods:
                    if p not in by_period:
                        claimed.append(None); avail.append(None); flags.append("OK")
                        continue
                    match = next((r for r in by_period[p]["rows"] if r["category"] == cat and r["tax_head"] == head), None)
                    if match:
                        claimed.append(match["itc_claimed_3b"])
                        avail.append(match["itc_available_2b"])
                        flags.append(match["flag"])
                    else:
                        claimed.append(None); avail.append(None); flags.append("OK")
                if all(v in (None, 0) for v in claimed) and all(v in (None, 0) for v in avail):
                    continue
                diff = [None if (a is None or b is None) else round(a - b, 2) for a, b in zip(claimed, avail)]
                row_flag = "CLAIMED > AVAILABLE" if "CLAIMED > AVAILABLE" in flags else ("REVIEW" if "REVIEW" in flags else "OK")
                _write_data_row(ws, row, gstin, f"R3B: {cat}", head, claimed, num_cols); row += 1
                _write_data_row(ws, row, gstin, f"R2B: {cat}", head, avail, num_cols); row += 1
                _write_data_row(ws, row, gstin, "Difference", head, diff, num_cols, italic=True, bold_label=True, flag=row_flag); row += 1

    _autofit(ws, len(periods))


TOLERANCE = 100  # rupees; matches comparator.py -- variances below this are rounding noise


def collect_action_items(gstr3b_records, gstr1_records, cmp_3b2b, periods):
    """Scans both comparison views and pulls out every flagged variance into a
    single flat list, sorted by absolute rupee materiality -- so instead of
    scrolling 6 tabs looking for problems, there's one ranked list to work from."""
    items = []
    r3b_by_gstin = _records_by_gstin_period(gstr3b_records)
    r1_by_gstin = _records_by_gstin_period(gstr1_records)
    gstins = sorted(set(r3b_by_gstin) | set(r1_by_gstin))

    for gstin in gstins:
        r3b_periods = r3b_by_gstin.get(gstin, {})
        r1_periods = r1_by_gstin.get(gstin, {})
        for section in _3B_VS_1_SECTIONS:
            for t in section["types"]:
                for p in periods:
                    if p not in r3b_periods:
                        continue
                    r3b_v = r3b_periods[p][section["r3b_key"]].get(t, 0) or 0
                    r1_sum = 0.0
                    any_r1 = False
                    for label, field, sign in section["components"]:
                        if p in r1_periods:
                            r1_sum += sign * (r1_periods[p][field].get(t, 0) or 0)
                            any_r1 = True
                    if not any_r1 and r3b_v == 0:
                        continue
                    diff = round(r3b_v - r1_sum, 2)
                    if abs(diff) > TOLERANCE:
                        items.append({
                            "gstin": gstin, "period": p, "source": "GSTR-3B vs GSTR-1",
                            "line_item": section["r3b_label"].replace("R3B: ", ""),
                            "tax_head": TYPE_LABELS[t], "variance": diff, "flag": "REVIEW",
                        })

    for c in cmp_3b2b:
        for row in c["rows"]:
            if row["flag"] == "OK":
                continue
            if row["category"] == "Reverse Charge (RCM)":
                # Expected structural gap: RCM/ISRC credit is self-declared against
                # cash-paid tax, not sourced from a supplier's GSTR-1 filing, so it
                # will almost always differ from GSTR-2B's revsup category. Flagging
                # this every month is noise, not a finding -- visible in the detail
                # sheet for context, but excluded from the actionable list.
                continue
            items.append({
                "gstin": c["gstin"], "period": c["period"], "source": "GSTR-3B vs GSTR-2B",
                "line_item": row["category"], "tax_head": row["tax_head"],
                "variance": row["variance"], "flag": row["flag"],
            })

    items.sort(key=lambda x: -abs(x["variance"]))
    return items


def _severity(abs_variance):
    if abs_variance > 100000:
        return "High"
    if abs_variance > 10000:
        return "Medium"
    return "Low"


def build_action_items_sheet(wb, items):
    ws = wb.create_sheet("Action Items", 2)  # right after Executive Summary
    headers = ["GSTIN", "Period", "Source", "Line Item", "Tax Head", "Variance (₹)", "Severity", "Flag"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = THIN_BORDER
    ws.freeze_panes = ws.cell(row=2, column=1)
    ws.auto_filter.ref = ws.dimensions if items else "A1:H1"

    if not items:
        c = ws.cell(row=2, column=1, value="No variances above tolerance (₹100) found across GSTR-3B vs GSTR-1 or GSTR-3B vs GSTR-2B.")
        c.font = Font(name=FONT_NAME, size=FONT_SIZE, italic=True, color="2E7D32")
        ws.column_dimensions["A"].width = 80
        return

    for i, item in enumerate(items, start=2):
        abs_var = abs(item["variance"])
        sev = _severity(abs_var)
        row_vals = [item["gstin"], period_label(item["period"]), item["source"], item["line_item"],
                    item["tax_head"], item["variance"], sev, item["flag"]]
        for col, v in enumerate(row_vals, start=1):
            c = ws.cell(row=i, column=col, value=v)
            c.font = BODY_FONT
            c.border = THIN_BORDER
            if col == 6:
                c.number_format = INR_FMT
        if sev in ("High", "Medium") or item["flag"] == "CLAIMED > AVAILABLE":
            for col in range(1, len(headers) + 1):
                ws.cell(row=i, column=col).fill = DIFF_FILL

    widths = [17, 12, 20, 42, 14, 16, 10, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_unmapped_data_sheet(wb, unmapped_info):
    """Surfaces any top-level JSON fields present in the source files that
    this tool doesn't actively parse into a polished section elsewhere --
    so nothing sits invisible just because we don't have a dedicated report
    section for it yet. A short preview only, not a full dump."""
    ws = wb.create_sheet("Unmapped Data")
    ws.sheet_properties.tabColor = "FFFFFF"
    headers = ["GSTIN", "Period", "Return Type", "Field Name", "Preview"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = THIN_BORDER
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:E{max(2, len(unmapped_info) + 1)}"

    if not unmapped_info:
        c = ws.cell(row=2, column=1, value="Every field in your uploaded files is reflected somewhere in this report.")
        c.font = Font(name=FONT_NAME, size=FONT_SIZE, italic=True, color="5B6472")
        ws.column_dimensions["A"].width = 80
        return

    for i, item in enumerate(sorted(unmapped_info, key=lambda x: (x["return_type"], x["period"], x["key"])), start=2):
        row_vals = [item["gstin"], period_label(item["period"]), item["return_type"], item["key"], item["preview"]]
        for col, v in enumerate(row_vals, start=1):
            c = ws.cell(row=i, column=col, value=v)
            c.font = BODY_FONT
            c.border = THIN_BORDER

    widths = [17, 12, 14, 20, 100]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_executive_summary_sheet(wb, gstr3b_records, gstr2b_detail_records, periods):
    """One-page headline view -- the handful of numbers that actually matter,
    before anyone has to go looking through the detailed sheets. Every figure
    here is already computed elsewhere; this just surfaces it in one place."""
    ws = wb.create_sheet("Executive Summary", 1)  # right after Overview
    ws.sheet_properties.tabColor = "FFFFFF"
    num_cols = _write_header(ws, periods)
    row = 2
    by_gstin_3b = _records_by_gstin_period(gstr3b_records)
    by_gstin_2b = _records_by_gstin_period(gstr2b_detail_records or [])

    def total_itc_available(r2b_detail):
        if not r2b_detail:
            return 0
        total = sum(r2b_detail["imports"].get(h, 0) for h in ("igst", "cgst", "sgst", "cess"))
        for cat_buckets in r2b_detail["by_category"].values():
            for elig_key in ("eligible_regular", "eligible_rcm"):
                total += sum(cat_buckets[elig_key].get(h, 0) for h in ("igst", "cgst", "sgst", "cess"))
        return total

    row_defs = [
        ("Total Outward Taxable Value", lambda r, r2b: r["total_liability_txval"]),
        ("Total Output Tax Liability", lambda r, r2b: r["total_liability_tax"]),
        ("Total RCM Liability", lambda r, r2b: sum(r["rcm_inward_liability"][h] for h in ("igst", "cgst", "sgst", "cess"))),
        ("Total ITC Available (GSTR-2B)", lambda r, r2b: total_itc_available(r2b)),
        ("Net ITC Claimed (GSTR-3B)", lambda r, r2b: sum(r["itc_net"][h] for h in ("igst", "cgst", "sgst"))),
        ("Tax Paid via Cash", lambda r, r2b: r["tax_paid_cash"]),
        ("Tax Paid via ITC", lambda r, r2b: r["tax_paid_itc"]),
        ("Interest Due", lambda r, r2b: sum(r["interest_due"].values())),
        ("Late Fee Due", lambda r, r2b: sum(r["late_fee_due"].values())),
    ]

    for gstin, by_period_3b in by_gstin_3b.items():
        by_period_2b = by_gstin_2b.get(gstin, {})
        for label, fn in row_defs:
            values = []
            for p in periods:
                if p not in by_period_3b:
                    values.append(None)
                    continue
                values.append(fn(by_period_3b[p], by_period_2b.get(p)))
            if all(v in (None, 0) for v in values):
                continue
            _write_data_row(ws, row, gstin, label, "Value", values, num_cols, bold_label=True)
            row += 1

    _autofit(ws, len(periods))


def build_workbook(gstr3b_records, gstr1_records, gstr2b_records, cmp_3b1, cmp_3b2b, output_path,
                    company_name=None, financial_year=None, data_notes=None, unmapped_info=None,
                    gstr2b_detail_records=None):
    periods = sorted_periods(gstr3b_records, gstr1_records, gstr2b_records)
    gstin = gstr3b_records[0]["gstin"] if gstr3b_records else (gstr1_records[0]["gstin"] if gstr1_records else None)

    wb = Workbook()
    wb.remove(wb.active)
    build_overview_sheet(wb, company_name, gstin, financial_year, data_notes=data_notes)
    build_executive_summary_sheet(wb, gstr3b_records, gstr2b_detail_records, periods)
    build_gstr3b_sheet(wb, gstr3b_records, periods)
    build_gstr1_sheet(wb, gstr1_records, periods)
    build_hsn_summary_sheet(wb, gstr1_records, periods)
    build_doc_issue_summary_sheet(wb, gstr1_records, periods)
    build_gstr2b_sheet(wb, gstr2b_detail_records or [], periods)
    build_3b_vs_1_sheet(wb, gstr3b_records, gstr1_records, periods)
    build_3b_vs_2b_sheet(wb, cmp_3b2b, periods)

    # Dedupe: the same unmapped key (e.g. 'qn') usually appears identically
    # every period -- show each distinct (return_type, key) combo once, not
    # once per month, to avoid a wall of repeated rows.
    seen = set()
    deduped_unmapped = []
    for item in (unmapped_info or []):
        dedupe_key = (item["return_type"], item["key"])
        if dedupe_key in seen:
            continue
        seen.add(dedupe_key)
        deduped_unmapped.append(item)
    build_unmapped_data_sheet(wb, deduped_unmapped)

    wb.save(output_path)
    return output_path
